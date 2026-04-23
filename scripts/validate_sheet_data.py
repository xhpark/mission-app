#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import pathlib
import sys
from typing import Iterable, Mapping

REQUIRED_COLUMNS = [
    "id",
    "category",
    "korean_text",
    "thai_text",
    "phonetic",
    "hangul_pronunciation",
]


def _read_rows(path: pathlib.Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "," if suffix == ".csv" else "\t"
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            return [
                {(k or "").strip(): (v or "").strip() for k, v in row.items()}
                for row in reader
            ]

    if suffix == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except Exception:
            raise RuntimeError("xlsx 검증을 위해 openpyxl 설치가 필요합니다. (pip install openpyxl)")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            mapped = {
                headers[i]: ("" if row[i] is None else str(row[i]).strip())
                for i in range(min(len(headers), len(row)))
            }
            if any(mapped.values()):
                rows.append(mapped)
        return rows

    raise RuntimeError(f"지원하지 않는 파일 형식: {suffix}")


def _missing_columns(headers: Iterable[str]) -> list[str]:
    normalized = {h.strip().lower() for h in headers if h}
    return [c for c in REQUIRED_COLUMNS if c not in normalized]


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate upload sheet structure.")
    parser.add_argument("sheet", type=pathlib.Path, help="Path to source xlsx/csv file")
    args = parser.parse_args()

    if not args.sheet.exists():
        print(f"[ERROR] file not found: {args.sheet}")
        return 1

    try:
        rows = _read_rows(args.sheet)
    except Exception as e:
        print(f"[ERROR] failed to read sheet: {e}")
        return 1

    if not rows:
        print("[ERROR] 데이터 행이 없습니다.")
        return 1

    missing = _missing_columns(rows[0].keys())
    if missing:
        print(f"[ERROR] 필수 컬럼 누락: {', '.join(missing)}")
        return 1

    id_set = set()
    for idx, row in enumerate(rows, start=2):
        item_id = row.get("id", "").strip()
        if not item_id:
            print(f"[ERROR] {idx}행: id 비어 있음")
            return 1
        if item_id in id_set:
            print(f"[ERROR] {idx}행: 중복 id ({item_id})")
            return 1
        id_set.add(item_id)

    print(f"[OK] 검증 통과: rows={len(rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
